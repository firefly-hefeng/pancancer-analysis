"""
单细胞RNA-seq元数据清洗与标准化工具
处理KPMP和ABC Atlas数据，识别并过滤非人源、非单细胞、非RNA-seq数据
使用Moonshot AI (Kimi)辅助模糊识别
"""

import pandas as pd
import numpy as np
import json
import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Set
import logging
from datetime import datetime
import requests
from concurrent.futures import ThreadPoolExecutor, as_completed
import time
from collections import defaultdict

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('data_cleaning.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


class KimiAIAssistant:
    """Moonshot AI (Kimi) 辅助识别器"""
    
    def __init__(self, api_keys: List[str]):
        self.api_keys = api_keys
        self.current_key_idx = 0
        self.base_url = "https://api.moonshot.cn/v1/chat/completions"
        self.request_count = defaultdict(int)
        self.max_requests_per_key = 50  # 每个key的请求限制
        
    def _get_next_key(self) -> str:
        """轮询获取下一个可用的API key"""
        for _ in range(len(self.api_keys)):
            key = self.api_keys[self.current_key_idx]
            if self.request_count[key] < self.max_requests_per_key:
                self.current_key_idx = (self.current_key_idx + 1) % len(self.api_keys)
                return key
            self.current_key_idx = (self.current_key_idx + 1) % len(self.api_keys)
        
        # 如果所有key都达到限制,等待并重置
        logger.warning("所有API key达到限制,等待60秒后重置...")
        time.sleep(60)
        self.request_count.clear()
        return self.api_keys[0]
    
    def identify_data_type(self, record: Dict) -> Dict[str, any]:
        """
        使用Kimi识别数据类型
        返回: {
            'is_human': bool,
            'is_single_cell': bool,
            'is_rna_seq': bool,
            'confidence': float,
            'reasoning': str
        }
        """
        api_key = self._get_next_key()
        
        # 构建prompt
        prompt = f"""
请分析以下生物医学数据集的元数据,判断它是否符合以下三个标准:
1. 是否为人类(Homo sapiens)数据?
2. 是否为单细胞/单核精度数据(scRNA-seq, snRNA-seq)?
3. 是否为RNA测序数据(而非ATAC-seq, ChIP-seq等)?

元数据信息:
- 标题: {record.get('title', 'N/A')}
- 实验设计: {record.get('experiment_design', 'N/A')}
- 样本类型: {record.get('sample_type', 'N/A')}
- 测序平台: {record.get('sequencing_platform', 'N/A')}
- 组织: {record.get('tissue', 'N/A')}
- 物种: {record.get('organism', 'N/A')}
- 摘要: {record.get('summary', 'N/A')[:500]}
- 补充信息: {str(record.get('supplementary_information', 'N/A'))[:300]}

请以JSON格式返回结果:
{{
    "is_human": true/false,
    "is_single_cell": true/false,
    "is_rna_seq": true/false,
    "confidence": 0.0-1.0,
    "reasoning": "简短说明理由"
}}
"""
        
        headers = {
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json"
        }
        
        data = {
            "model": "moonshot-v1-8k",
            "messages": [
                {"role": "system", "content": "你是一个生物信息学数据分析专家,擅长识别测序数据的类型和质量。请始终以JSON格式返回结果。"},
                {"role": "user", "content": prompt}
            ],
            "temperature": 0.3,
            "max_tokens": 500
        }
        
        try:
            response = requests.post(self.base_url, headers=headers, json=data, timeout=30)
            response.raise_for_status()
            
            result = response.json()
            content = result['choices'][0]['message']['content']
            
            # 提取JSON
            json_match = re.search(r'\{[^{}]*\}', content, re.DOTALL)
            if json_match:
                analysis = json.loads(json_match.group())
                self.request_count[api_key] += 1
                return analysis
            else:
                logger.warning(f"无法解析Kimi响应: {content}")
                return self._default_response()
                
        except Exception as e:
            logger.error(f"Kimi API调用失败: {e}")
            return self._default_response()
    
    def _default_response(self) -> Dict:
        """默认响应(调用失败时)"""
        return {
            'is_human': None,
            'is_single_cell': None,
            'is_rna_seq': None,
            'confidence': 0.0,
            'reasoning': 'API调用失败,使用规则判断'
        }


class DataCleaner:
    """数据清洗器"""
    
    def __init__(self, input_dir: str, output_dir: str, use_kimi: bool = True):
        self.input_dir = Path(input_dir)
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        self.use_kimi = use_kimi
        if use_kimi:
            api_keys = [
                "sk-1zQD2QV94Vi8p6iCifYoqEFqWJnqSF4BICv4wOz8C5aG3bMy"
            ]
            self.kimi = KimiAIAssistant(api_keys)
        
        # 规则库
        self.human_keywords = [
            'homo sapiens', 'human', 'patient', 'donor', 'subject',
            'hg19', 'hg38', 'grch37', 'grch38'
        ]
        
        self.non_human_keywords = [
            'mouse', 'mus musculus', 'rat', 'rattus', 'drosophila',
            'zebrafish', 'danio rerio', 'c. elegans', 'yeast',
            'arabidopsis', 'mm10', 'mm39', 'rn6'
        ]
        
        self.single_cell_keywords = [
            'scrna-seq', 'scrna', 'snrna-seq', 'snrna', 'single cell',
            'single-cell', 'single nucleus', 'single-nucleus',
            'droplet', '10x', '10x genomics', 'drop-seq', 'smart-seq',
            'cel-seq', 'mars-seq', 'indrops'
        ]
        
        self.bulk_keywords = [
            'bulk rna-seq', 'bulk rna', 'bulk sequencing',
            'population', 'tissue-level', 'whole tissue'
        ]
        
        self.rna_seq_keywords = [
            'rna-seq', 'rna seq', 'transcriptome', 'gene expression',
            'mrna', 'transcript'
        ]
        
        self.non_rna_keywords = [
            'atac-seq', 'atac', 'chip-seq', 'chip', 'wgs', 'wes',
            'whole genome', 'whole exome', 'dna-seq', 'methylation',
            'bisulfite', 'cut&run', 'cut&tag', 'dnase-seq',
            'hi-c', 'chromatin', 'accessibility'
        ]
        
        # 统计信息
        self.stats = {
            'total': 0,
            'passed': 0,
            'failed_non_human': 0,
            'failed_bulk': 0,
            'failed_non_rna': 0,
            'uncertain': 0,
            'kimi_used': 0
        }
    
    def load_data(self, database: str) -> pd.DataFrame:
        """加载数据"""
        logger.info(f"加载 {database} 数据...")
        
        if database.lower() == 'kpmp':
            file_pattern = 'kpmp_series_metadata.csv'
        elif database.lower() == 'abc':
            file_pattern = 'FINAL_abc_atlas*.csv'
        else:
            raise ValueError(f"不支持的数据库: {database}")
        
        # 查找文件
        files = list(self.input_dir.glob(f"**/{file_pattern}"))
        
        if not files:
            logger.error(f"未找到匹配的文件: {file_pattern}")
            return pd.DataFrame()
        
        logger.info(f"找到文件: {files[0]}")
        df = pd.read_csv(files[0])
        logger.info(f"加载了 {len(df)} 条记录")
        
        return df
    
    def rule_based_check(self, record: Dict) -> Tuple[bool, bool, bool, str]:
        """
        基于规则的检查
        返回: (is_human, is_single_cell, is_rna_seq, reason)
        """
        text = ' '.join([
            str(record.get('title', '')),
            str(record.get('experiment_design', '')),
            str(record.get('sample_type', '')),
            str(record.get('sequencing_platform', '')),
            str(record.get('tissue', '')),
            str(record.get('organism', '')),
            str(record.get('summary', '')),
            str(record.get('supplementary_information', ''))
        ]).lower()
        
        reasons = []
        
        # 1. 人类检查
        is_human = None
        if any(kw in text for kw in self.non_human_keywords):
            is_human = False
            reasons.append("检测到非人类物种关键词")
        elif any(kw in text for kw in self.human_keywords):
            is_human = True
            reasons.append("检测到人类关键词")
        
        # 2. 单细胞检查
        is_single_cell = None
        if any(kw in text for kw in self.bulk_keywords):
            is_single_cell = False
            reasons.append("检测到bulk测序关键词")
        elif any(kw in text for kw in self.single_cell_keywords):
            is_single_cell = True
            reasons.append("检测到单细胞关键词")
        
        # 3. RNA-seq检查
        is_rna_seq = None
        if any(kw in text for kw in self.non_rna_keywords):
            is_rna_seq = False
            reasons.append("检测到非RNA测序技术")
        elif any(kw in text for kw in self.rna_seq_keywords):
            is_rna_seq = True
            reasons.append("检测到RNA测序关键词")
        
        reason = '; '.join(reasons) if reasons else '无明确关键词'
        
        return is_human, is_single_cell, is_rna_seq, reason
    
    def check_record(self, record: Dict, use_kimi_for_uncertain: bool = True) -> Dict:
        """
        检查单条记录
        """
        # 首先使用规则检查
        is_human_rule, is_sc_rule, is_rna_rule, rule_reason = self.rule_based_check(record)
        
        # 判断是否需要Kimi辅助
        need_kimi = (
            use_kimi_for_uncertain and 
            self.use_kimi and
            (is_human_rule is None or is_sc_rule is None or is_rna_rule is None)
        )
        
        if need_kimi:
            logger.info(f"使用Kimi辅助识别: {record.get('id', 'unknown')}")
            kimi_result = self.kimi.identify_data_type(record)
            self.stats['kimi_used'] += 1
            
            # 合并结果(Kimi优先)
            is_human = kimi_result.get('is_human') if kimi_result.get('is_human') is not None else is_human_rule
            is_sc = kimi_result.get('is_single_cell') if kimi_result.get('is_single_cell') is not None else is_sc_rule
            is_rna = kimi_result.get('is_rna_seq') if kimi_result.get('is_rna_seq') is not None else is_rna_rule
            
            reason = f"规则: {rule_reason}; Kimi: {kimi_result.get('reasoning', '')}"
            confidence = kimi_result.get('confidence', 0.5)
        else:
            is_human = is_human_rule
            is_sc = is_sc_rule
            is_rna = is_rna_rule
            reason = f"规则判断: {rule_reason}"
            confidence = 0.8 if all(x is not None for x in [is_human, is_sc, is_rna]) else 0.3
        
        # 判断是否通过
        passed = (is_human == True) and (is_sc == True) and (is_rna == True)
        
        # 判断失败原因
        if not passed:
            if is_human == False:
                self.stats['failed_non_human'] += 1
                failure_reason = 'non_human'
            elif is_sc == False:
                self.stats['failed_bulk'] += 1
                failure_reason = 'bulk_or_non_single_cell'
            elif is_rna == False:
                self.stats['failed_non_rna'] += 1
                failure_reason = 'non_rna_seq'
            else:
                self.stats['uncertain'] += 1
                failure_reason = 'uncertain'
        else:
            self.stats['passed'] += 1
            failure_reason = None
        
        return {
            'passed': passed,
            'is_human': is_human,
            'is_single_cell': is_sc,
            'is_rna_seq': is_rna,
            'confidence': confidence,
            'reason': reason,
            'failure_reason': failure_reason
        }
    
    def extract_ids(self, record: Dict) -> Dict[str, str]:
        """
        提取各类ID
        """
        ids = {
            'Study/Project_id_1': '',
            'Study/Project_id_2': '',
            'Study/Project_id_3': '',
            'raw_sample_id': '',
            'matrix_sample_id': ''
        }
        
        # 从source_database和id提取
        source_db = str(record.get('source_database', '')).lower()
        main_id = str(record.get('id', ''))
        
        # 从supplementary_information提取额外ID
        supp_info = str(record.get('supplementary_information', ''))
        
        # 提取GEO ID
        geo_match = re.search(r'(GSE\d+)', main_id + ' ' + supp_info, re.IGNORECASE)
        if geo_match:
            ids['Study/Project_id_1'] = geo_match.group(1)
        
        # 提取SRA ID
        sra_match = re.search(r'(SRP\d+|ERP\d+|DRP\d+)', supp_info, re.IGNORECASE)
        if sra_match:
            if ids['Study/Project_id_1']:
                ids['Study/Project_id_2'] = sra_match.group(1)
            else:
                ids['Study/Project_id_1'] = sra_match.group(1)
        
        # 提取ArrayExpress ID
        ae_match = re.search(r'(E-[A-Z]+-\d+)', supp_info, re.IGNORECASE)
        if ae_match:
            if ids['Study/Project_id_1'] and ids['Study/Project_id_2']:
                ids['Study/Project_id_3'] = ae_match.group(1)
            elif ids['Study/Project_id_1']:
                ids['Study/Project_id_2'] = ae_match.group(1)
            else:
                ids['Study/Project_id_1'] = ae_match.group(1)
        
        # 如果没有提取到标准ID,使用原始ID
        if not ids['Study/Project_id_1']:
            ids['Study/Project_id_1'] = main_id
        
        # raw vs matrix sample ID判断
        if 'raw' in source_db or 'sra' in source_db or 'repository' in source_db:
            ids['raw_sample_id'] = main_id
        elif 'geo' in source_db or 'matrix' in source_db or 'cellxgene' in source_db:
            ids['matrix_sample_id'] = main_id
        else:
            # 默认都填
            ids['raw_sample_id'] = main_id
            ids['matrix_sample_id'] = main_id
        
        return ids
    
    def determine_data_availability(self, record: Dict) -> Dict[str, any]:
        """
        判断数据可用性
        """
        availability = {
            'raw_exist': False,
            'raw_open': False,
            'matrix_exist': False,
            'matrix_open': False,
            'file_type': ''
        }
        
        source_db = str(record.get('source_database', '')).lower()
        open_status = str(record.get('open_status', '')).lower()
        data_tier = str(record.get('data_tier', '')).lower()
        supp_info = str(record.get('supplementary_information', '')).lower()
        
        # 判断raw数据
        if any(kw in source_db for kw in ['sra', 'repository', 'raw']):
            availability['raw_exist'] = True
        if 'raw' in data_tier or 'fastq' in supp_info:
            availability['raw_exist'] = True
        
        # 判断matrix数据
        if any(kw in source_db for kw in ['geo', 'cellxgene', 'matrix', 'processed']):
            availability['matrix_exist'] = True
        if 'processed' in data_tier or 'matrix' in data_tier:
            availability['matrix_exist'] = True
        
        # 判断开放状态
        is_open = 'public' in open_status or 'open' in open_status
        
        if availability['raw_exist']:
            availability['raw_open'] = is_open
        if availability['matrix_exist']:
            availability['matrix_open'] = is_open
        
        # 文件类型
        if 'fastq' in supp_info or 'raw' in data_tier:
            availability['file_type'] = 'FASTQ/BAM'
        elif 'h5ad' in supp_info or 'cellxgene' in source_db:
            availability['file_type'] = 'H5AD'
        elif 'matrix' in data_tier or 'mtx' in supp_info:
            availability['file_type'] = 'Matrix'
        else:
            availability['file_type'] = 'Unknown'
        
        return availability
    
    def standardize_sample_type(self, record: Dict) -> str:
        """
        标准化sample_type为: healthy, tumor, other_illness
        """
        disease = str(record.get('disease', '')).lower()
        disease_general = str(record.get('disease_general', '')).lower()
        
        if any(kw in disease or kw in disease_general for kw in ['normal', 'healthy', 'control']):
            return 'healthy'
        elif any(kw in disease or kw in disease_general for kw in ['cancer', 'tumor', 'carcinoma', 'malignant']):
            return 'tumor'
        else:
            return 'other_illness'
    
    def transform_to_standard_format(self, record: Dict) -> Dict[str, any]:
        """
        转换为标准格式
        """
        # 提取ID
        ids = self.extract_ids(record)
        
        # 判断数据可用性
        availability = self.determine_data_availability(record)
        
        # 标准化sample_type
        sample_type = self.standardize_sample_type(record)
        
        standard_record = {
            'Study/Project_id_1': ids['Study/Project_id_1'],
            'Study/Project_id_2': ids['Study/Project_id_2'],
            'Study/Project_id_3': ids['Study/Project_id_3'],
            'raw_sample_id': ids['raw_sample_id'],
            'matrix_sample_id': ids['matrix_sample_id'],
            'raw_exist': availability['raw_exist'],
            'raw_open': availability['raw_open'],
            'matrix_exist': availability['matrix_exist'],
            'matrix_open': availability['matrix_open'],
            'file_type': availability['file_type'],
            'title': record.get('title', ''),
            'disease_general': record.get('disease_general', ''),
            'disease': record.get('disease', ''),
            'pubmed': record.get('pubmed', ''),
            'source_database': record.get('source_database', ''),
            'access_link': record.get('access_link', ''),
            'open_status': record.get('open_status', ''),
            'ethnicity': record.get('ethnicity', ''),
            'sex': record.get('sex', ''),
            'tissue_location': record.get('tissue_location', record.get('tissue', '')),
            'sequencing_platform': record.get('sequencing_platform', ''),
            'experiment_design': record.get('experiment_design', ''),
            'sample_type': sample_type,
            'summary': record.get('summary', ''),
            'citation_count': record.get('citation_count', ''),
            'publication_date': record.get('publication_date', ''),
            'submission_date': record.get('submission_date', ''),
            'last_update_date': record.get('last_update_date', ''),
            'contact_name': record.get('contact_name', ''),
            'contact_email': record.get('contact_email', ''),
            'contact_institute': record.get('contact_institute', ''),
            'supplementary_information': record.get('supplementary_information', '')
        }
        
        return standard_record
    
    def process_database(self, database: str) -> Tuple[pd.DataFrame, pd.DataFrame]:
        """
        处理单个数据库
        返回: (通过的数据, 混杂的数据)
        """
        logger.info(f"\n{'='*80}")
        logger.info(f"处理 {database} 数据库")
        logger.info(f"{'='*80}\n")
        
        # 加载数据
        df = self.load_data(database)
        
        if df.empty:
            logger.warning(f"{database} 数据为空")
            return pd.DataFrame(), pd.DataFrame()
        
        self.stats['total'] += len(df)
        
        # 处理每条记录
        passed_records = []
        flagged_records = []
        
        for idx, row in df.iterrows():
            if idx % 10 == 0:
                logger.info(f"处理进度: {idx+1}/{len(df)}")
            
            record = row.to_dict()
            
            # 检查记录
            check_result = self.check_record(record, use_kimi_for_uncertain=True)
            
            # 转换为标准格式
            standard_record = self.transform_to_standard_format(record)
            
            # 添加检查信息
            standard_record['check_is_human'] = check_result['is_human']
            standard_record['check_is_single_cell'] = check_result['is_single_cell']
            standard_record['check_is_rna_seq'] = check_result['is_rna_seq']
            standard_record['check_confidence'] = check_result['confidence']
            standard_record['check_reason'] = check_result['reason']
            
            if check_result['passed']:
                # 移除检查字段后加入通过列表
                clean_record = {k: v for k, v in standard_record.items() 
                               if not k.startswith('check_')}
                passed_records.append(clean_record)
            else:
                # 保留检查字段,加入标记列表
                standard_record['failure_reason'] = check_result['failure_reason']
                flagged_records.append(standard_record)
            
            # 避免API限流
            if self.use_kimi and (idx + 1) % 20 == 0:
                logger.info("暂停3秒避免API限流...")
                time.sleep(3)
        
        passed_df = pd.DataFrame(passed_records)
        flagged_df = pd.DataFrame(flagged_records)
        
        logger.info(f"\n{database} 处理完成:")
        logger.info(f"  总记录数: {len(df)}")
        logger.info(f"  通过: {len(passed_df)}")
        logger.info(f"  标记(需人工复核): {len(flagged_df)}")
        
        return passed_df, flagged_df
    
    def merge_duplicates(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        合并重复记录(同一项目在不同数据库中)
        """
        logger.info("\n合并重复记录...")
        
        if df.empty:
            return df
        
        # 基于标题和主要ID分组
        df['merge_key'] = df.apply(
            lambda x: f"{x['title']}_{x['Study/Project_id_1']}", 
            axis=1
        )
        
        merged_records = []
        
        for merge_key, group in df.groupby('merge_key'):
            if len(group) == 1:
                merged_records.append(group.iloc[0].to_dict())
            else:
                # 合并多条记录
                merged = self._merge_record_group(group)
                merged_records.append(merged)
        
        merged_df = pd.DataFrame(merged_records)
        
        # 删除临时列
        if 'merge_key' in merged_df.columns:
            merged_df = merged_df.drop('merge_key', axis=1)
        
        logger.info(f"  合并前: {len(df)} 条")
        logger.info(f"  合并后: {len(merged_df)} 条")
        
        return merged_df
    
    def _merge_record_group(self, group: pd.DataFrame) -> Dict:
        """合并一组重复记录"""
        merged = group.iloc[0].to_dict()
        
        # 合并ID字段
        all_ids = set()
        for idx, row in group.iterrows():
            for id_col in ['Study/Project_id_1', 'Study/Project_id_2', 'Study/Project_id_3']:
                if pd.notna(row[id_col]) and row[id_col]:
                    all_ids.add(str(row[id_col]))
        
        id_list = sorted(list(all_ids))
        for i in range(min(3, len(id_list))):
            merged[f'Study/Project_id_{i+1}'] = id_list[i]
        
        # 合并raw和matrix ID
        raw_ids = group['raw_sample_id'].dropna().unique()
        matrix_ids = group['matrix_sample_id'].dropna().unique()
        
        if len(raw_ids) > 0:
            merged['raw_sample_id'] = '; '.join(str(x) for x in raw_ids)
        if len(matrix_ids) > 0:
            merged['matrix_sample_id'] = '; '.join(str(x) for x in matrix_ids)
        
        # 数据可用性: 任一为True则为True
        merged['raw_exist'] = group['raw_exist'].any()
        merged['raw_open'] = group['raw_open'].any()
        merged['matrix_exist'] = group['matrix_exist'].any()
        merged['matrix_open'] = group['matrix_open'].any()
        
        # 合并来源数据库
        sources = group['source_database'].dropna().unique()
        merged['source_database'] = '; '.join(str(x) for x in sources)
        
        # 其他字段取第一个非空值
        for col in merged.keys():
            if pd.isna(merged[col]) or merged[col] == '':
                non_null = group[col].dropna()
                if len(non_null) > 0:
                    merged[col] = non_null.iloc[0]
        
        return merged
    
    def save_results(self, passed_df: pd.DataFrame, flagged_df: pd.DataFrame, 
                    database: str):
        """保存结果"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # 保存通过的数据
        if not passed_df.empty:
            passed_file = self.output_dir / f"cleaned_{database}_passed_{timestamp}.csv"
            passed_df.to_csv(passed_file, index=False, encoding='utf-8-sig')
            logger.info(f"✓ 已保存通过的数据: {passed_file}")
            
            # 保存Excel格式
            excel_file = self.output_dir / f"cleaned_{database}_passed_{timestamp}.xlsx"
            with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                passed_df.to_excel(writer, sheet_name='Cleaned_Data', index=False)
            logger.info(f"✓ 已保存Excel格式: {excel_file}")
        
        # 保存需要人工复核的数据
        if not flagged_df.empty:
            flagged_file = self.output_dir / f"flagged_{database}_for_review_{timestamp}.csv"
            flagged_df.to_csv(flagged_file, index=False, encoding='utf-8-sig')
            logger.info(f"⚠ 已保存待复核数据: {flagged_file}")
            
            # 保存Excel格式(带颜色标记)
            excel_file = self.output_dir / f"flagged_{database}_for_review_{timestamp}.xlsx"
            self._save_flagged_excel(flagged_df, excel_file)
    
    def _save_flagged_excel(self, df: pd.DataFrame, filepath: Path):
        """保存带颜色标记的Excel"""
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill
        
        # 先保存基本Excel
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Flagged_Data', index=False)
        
        # 加载并添加颜色
        wb = load_workbook(filepath)
        ws = wb['Flagged_Data']
        
        # 定义颜色
        colors = {
            'non_human': PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid'),
            'bulk_or_non_single_cell': PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid'),
            'non_rna_seq': PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid'),
            'uncertain': PatternFill(start_color='CCCCFF', end_color='CCCCFF', fill_type='solid')
        }
        
        # 找到failure_reason列
        failure_col_idx = None
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == 'failure_reason':
                failure_col_idx = idx
                break
        
        if failure_col_idx:
            for row_idx in range(2, ws.max_row + 1):
                reason = ws.cell(row_idx, failure_col_idx).value
                if reason in colors:
                    for col_idx in range(1, ws.max_column + 1):
                        ws.cell(row_idx, col_idx).fill = colors[reason]
        
        wb.save(filepath)
        logger.info(f"✓ 已保存带颜色标记的Excel: {filepath}")
    
    def generate_report(self, output_file: str = None):
        """生成清洗报告"""
        if output_file is None:
            output_file = self.output_dir / f"cleaning_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write("="*80 + "\n")
            f.write("数据清洗报告\n")
            f.write("="*80 + "\n\n")
            
            f.write(f"清洗时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
            
            f.write("统计信息:\n")
            f.write(f"  总记录数: {self.stats['total']}\n")
            f.write(f"  通过数量: {self.stats['passed']} ({self.stats['passed']/self.stats['total']*100:.1f}%)\n")
            f.write(f"  失败数量: {self.stats['total'] - self.stats['passed']}\n")
            f.write(f"    - 非人类: {self.stats['failed_non_human']}\n")
            f.write(f"    - Bulk/非单细胞: {self.stats['failed_bulk']}\n")
            f.write(f"    - 非RNA测序: {self.stats['failed_non_rna']}\n")
            f.write(f"    - 不确定: {self.stats['uncertain']}\n")
            f.write(f"  Kimi辅助识别次数: {self.stats['kimi_used']}\n\n")
            
            f.write("="*80 + "\n")
        
        logger.info(f"✓ 清洗报告已保存: {output_file}")
    
    def run(self, databases: List[str] = ['kpmp', 'abc']):
        """运行完整的清洗流程"""
        logger.info("\n" + "="*80)
        logger.info("开始数据清洗流程")
        logger.info("="*80 + "\n")
        
        all_passed = []
        all_flagged = []
        
        for db in databases:
            try:
                passed_df, flagged_df = self.process_database(db)
                
                if not passed_df.empty:
                    all_passed.append(passed_df)
                if not flagged_df.empty:
                    all_flagged.append(flagged_df)
                
                # 保存单个数据库结果
                self.save_results(passed_df, flagged_df, db)
                
            except Exception as e:
                logger.error(f"处理 {db} 时出错: {e}")
                import traceback
                traceback.print_exc()
        
        # 合并所有数据库
        if all_passed:
            logger.info("\n合并所有数据库的通过数据...")
            combined_passed = pd.concat(all_passed, ignore_index=True)
            
            # 去重合并
            combined_passed = self.merge_duplicates(combined_passed)
            
            # 保存最终结果
            final_file = self.output_dir / f"FINAL_cleaned_all_databases_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            combined_passed.to_csv(final_file, index=False, encoding='utf-8-sig')
            logger.info(f"✓ 最终清洗数据已保存: {final_file}")
            
            # Excel格式
            final_excel = self.output_dir / f"FINAL_cleaned_all_databases_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            with pd.ExcelWriter(final_excel, engine='openpyxl') as writer:
                combined_passed.to_excel(writer, sheet_name='Final_Cleaned', index=False)
            logger.info(f"✓ 最终Excel已保存: {final_excel}")
        
        if all_flagged:
            combined_flagged = pd.concat(all_flagged, ignore_index=True)
            flagged_file = self.output_dir / f"FINAL_flagged_all_databases_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            combined_flagged.to_csv(flagged_file, index=False, encoding='utf-8-sig')
            logger.info(f"⚠ 最终待复核数据已保存: {flagged_file}")
        
        # 生成报告
        self.generate_report()
        
        logger.info("\n" + "="*80)
        logger.info("数据清洗完成!")
        logger.info(f"输出目录: {self.output_dir}")
        logger.info("="*80 + "\n")


def main():
    """主程序"""
    print("\n" + "="*80)
    print("单细胞RNA-seq元数据清洗工具")
    print("="*80)
    print("\n功能:")
    print("  1. 识别并过滤非人源数据")
    print("  2. 识别并过滤Bulk测序数据")
    print("  3. 识别并过滤非RNA测序数据")
    print("  4. 使用Moonshot AI辅助模糊识别")
    print("  5. 标准化数据格式")
    print("  6. 合并重复记录")
    print("\n" + "="*80 + "\n")
    
    # 配置路径
    # 修改这些路径为你的实际路径
    kpmp_input_dir = "./kpmp_metadata_collection"
    abc_input_dir = "./abc_atlas_metadata"
    output_dir = "./cleaned_metadata"
    
    # 创建清洗器
    cleaner = DataCleaner(
        input_dir="./",  # 会在子目录中搜索
        output_dir=output_dir,
        use_kimi=True  # 启用Kimi辅助识别
    )
    
    # 手动设置输入目录
    cleaner.kpmp_dir = Path(kpmp_input_dir)
    cleaner.abc_dir = Path(abc_input_dir)
    
    # 运行清洗
    try:
        cleaner.run(databases=['kpmp', 'abc'])
        
        print("\n" + "="*80)
        print("清洗统计:")
        print(f"  总记录: {cleaner.stats['total']}")
        print(f"  通过: {cleaner.stats['passed']}")
        print(f"  标记: {cleaner.stats['total'] - cleaner.stats['passed']}")
        print(f"  Kimi调用: {cleaner.stats['kimi_used']} 次")
        print("="*80 + "\n")
        
    except Exception as e:
        logger.error(f"清洗过程出错: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()